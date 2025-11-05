# src/aura_core/apprentices/spreadsheet_creator.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import traceback
from datetime import datetime

class SpreadsheetCreator:
    def __init__(self, filename):
        if not filename.lower().endswith('.xlsx'):
            filename += '.xlsx'
        self.filename = filename
        self.workbook = openpyxl.Workbook()
        self.named_styles = {}
        # Remove the default 'Sheet'
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])

    def _apply_style(self, cell, style_dict):
        """Applies a style dictionary to a cell."""
        if not style_dict:
            return

        # Font
        font_attrs = {}
        if "font" in style_dict:
            if style_dict["font"].get("bold"): font_attrs["bold"] = True
            if style_dict["font"].get("italic"): font_attrs["italic"] = True
            if style_dict["font"].get("color"): font_attrs["color"] = style_dict["font"]["color"]
            if style_dict["font"].get("name"): font_attrs["name"] = style_dict["font"]["name"]
            if style_dict["font"].get("size"): font_attrs["size"] = float(style_dict["font"]["size"])
        if font_attrs:
            cell.font = Font(**font_attrs)
        
        # Fill
        if "fill" in style_dict and style_dict["fill"].get("color"):
            cell.fill = PatternFill(start_color=style_dict["fill"]["color"],
                                     end_color=style_dict["fill"]["color"],
                                     fill_type="solid")
        
        # Alignment
        align_attrs = {}
        if "align" in style_dict:
            align = style_dict["align"]
            h_map = {"left": "left", "center": "center", "right": "right", "justify": "justify"}
            v_map = {"top": "top", "center": "center", "bottom": "bottom"}
            if align.get("horizontal"): align_attrs["horizontal"] = h_map.get(align["horizontal"])
            if align.get("vertical"): align_attrs["vertical"] = v_map.get(align["vertical"])
            if align.get("wrap_text"): align_attrs["wrap_text"] = True
        if align_attrs:
            cell.alignment = Alignment(**align_attrs)

        # Number Format
        if "number_format" in style_dict:
            cell.number_format = style_dict["number_format"] # e.g., "$#,##0.00", "0.00%", "mm-dd-yy"

    def auto_fit_columns(self, sheet):
        """Auto-fit all columns in the given sheet."""
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter # Get the column letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

    def build(self, payload):
        """Builds the workbook from the payload."""
        # 1. Define named styles
        if "named_styles" in payload:
            for style_name, style_dict in payload["named_styles"].items():
                # In a real scenario, you'd convert this to openpyxl objects
                # For simplicity, we just store the dicts
                self.named_styles[style_name] = style_dict

        # 2. Process sheets
        sheets_data = payload.get("sheets", [])
        if not sheets_data and "data" in payload: # Backwards compatibility
            print("--- Spreadsheet Creator Info: Using legacy 'data' key. Converting to single sheet. ---")
            sheets_data = [{"name": "Sheet1", "data": payload.get("data")}]
        
        if not sheets_data:
             print("--- Spreadsheet Creator Warning: No sheets defined in payload. Creating empty file. ---")

        for i, sheet_def in enumerate(sheets_data):
            sheet_name = sheet_def.get("name", f"Sheet{i+1}")
            sheet = self.workbook.create_sheet(title=sheet_name)
            sheet_data = sheet_def.get("data", [])

            # Write data
            for r_idx, row_data in enumerate(sheet_data, start=1):
                for c_idx, cell_data in enumerate(row_data, start=1):
                    cell = sheet.cell(row=r_idx, column=c_idx)
                    
                    if isinstance(cell_data, dict):
                        # It's a complex cell definition
                        value = cell_data.get("value")
                        
                        if cell_data.get("is_formula"):
                            cell.value = str(value) # Write as formula, e.g., "=SUM(A1:B1)"
                        else:
                            cell.value = value # Write as regular value

                        # Apply styles
                        style_name = cell_data.get("style_name")
                        inline_style = cell_data.get("style")
                        if style_name and style_name in self.named_styles:
                            self._apply_style(cell, self.named_styles[style_name])
                        if inline_style:
                            self._apply_style(cell, inline_style)
                    
                    else:
                        # It's a simple value
                        cell.value = cell_data

            # Apply column widths
            if "column_widths" in sheet_def:
                if sheet_def["column_widths"] == "auto":
                    self.auto_fit_columns(sheet)
                elif isinstance(sheet_def["column_widths"], dict):
                    for col_letter, width in sheet_def["column_widths"].items():
                        try:
                            sheet.column_dimensions[col_letter].width = float(width)
                        except Exception as e:
                            print(f"--- Spreadsheet Creator Warning: Invalid width for col {col_letter}. Reason: {e} ---")
            
            # Freeze panes
            if "freeze_panes" in sheet_def:
                sheet.freeze_panes = sheet_def["freeze_panes"] # e.g., "B2"

    def save(self):
        """Saves the workbook."""
        xls_dir = os.path.dirname(self.filename)
        if xls_dir and not os.path.exists(xls_dir):
            try:
                os.makedirs(xls_dir, exist_ok=True)
                print(f"--- Spreadsheet Creator: Created destination directory '{xls_dir}' ---")
            except Exception as e:
                raise Exception(f"Could not create destination directory '{xls_dir}'. Reason: {e}")
        
        self.workbook.save(self.filename)


def run(payload):
    """
    Main entry point for the Foreman.
    Parses the payload, creates a SpreadsheetCreator instance, and builds the document.
    """
    try:
        filename = payload.get("filename")
        if not filename:
            return "Error: Missing 'filename' in payload."
            
        creator = SpreadsheetCreator(filename)
        creator.build(payload) # Pass the whole payload
        creator.save()
        
        return f"Success: Created spreadsheet '{filename}'."

    except Exception as e:
        print(traceback.format_exc()) # Print detailed error
        return f"Error creating spreadsheet '{filename}'. Reason: {e}"