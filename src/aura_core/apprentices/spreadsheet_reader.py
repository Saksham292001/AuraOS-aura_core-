# src/aura_core/apprentices/spreadsheet_reader.py
import openpyxl
import os
import traceback
from openpyxl.utils import rows_from_range, get_column_letter

def _get_sheet(workbook, sheet_name=None):
    """Safely gets a worksheet by name or defaults to active."""
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
        return workbook[sheet_name]
    else:
        print(f"--- Spreadsheet Reader Info: No sheet specified. Reading active sheet. ---")
        return workbook.active

def _read_as_list(sheet, read_range=None):
    """Reads data from a sheet/range into a list of lists."""
    data = []
    
    # Determine rows to iterate over
    rows_iter = sheet[read_range] if read_range else sheet.iter_rows()
    
    for row in rows_iter:
        row_data = [cell.value for cell in row]
        if any(cell_data is not None for cell_data in row_data): # Filter empty rows
            data.append(row_data)
    return data

def _read_as_dict(sheet, read_range=None):
    """
    Reads data from a sheet/range into a list of dictionaries.
    Assumes first non-empty row in range is the header.
    """
    data = []
    
    rows_iter = sheet[read_range] if read_range else sheet.iter_rows()
    
    # Find the header row
    header = None
    row_iterator = iter(rows_iter)
    
    while header is None:
        try:
            first_row = next(row_iterator)
            # Check if it's a real header (not just empty cells)
            if any(cell.value is not None for cell in first_row):
                header = [str(cell.value) for cell in first_row]
        except StopIteration:
            # Sheet is empty
            return [] 

    # Read data rows
    for row in row_iterator:
        row_data = [cell.value for cell in row]
        if any(cell_data is not None for cell_data in row_data):
            row_dict = {header[i]: row_data[i] for i in range(min(len(header), len(row_data)))}
            data.append(row_dict)
    return data


def run(payload):
    """
    Reads data from a specified Excel sheet (.xlsx).
    Can return list of lists, list of dicts, or all sheets.
    """
    filename = payload.get("filename")
    sheet_name = payload.get("sheet_name") # Specific name, "all", or None (active)
    read_mode = payload.get("read_mode", "list").lower() # "list" or "dict"
    read_range = payload.get("read_range") # e.g., "A1:D20"

    if not filename:
        return "Error: Missing 'filename' for the spreadsheet reader."
    if not os.path.exists(filename):
        return f"Error: File not found at '{filename}'."

    try:
        workbook = openpyxl.load_workbook(filename, data_only=True) # data_only=True reads formula results
        
        # Handle "all" sheets request
        if sheet_name and sheet_name.lower() == "all":
            all_data = {}
            for name in workbook.sheetnames:
                sheet = workbook[name]
                if read_mode == "dict":
                    all_data[name] = _read_as_dict(sheet, read_range=None) # Range ignored for "all"
                else:
                    all_data[name] = _read_as_list(sheet, read_range=None) # Range ignored for "all"
            workbook.close()
            return all_data

        # Handle single sheet request
        sheet = _get_sheet(workbook, sheet_name)
        
        if read_mode == "dict":
            data = _read_as_dict(sheet, read_range)
        else:
            data = _read_as_list(sheet, read_range)
            
        workbook.close()
        return data

    except Exception as e:
        if 'workbook' in locals() and workbook:
            workbook.close()
        print(traceback.format_exc())
        return f"Error reading spreadsheet '{filename}'. Reason: {e}"